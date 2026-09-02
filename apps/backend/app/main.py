import csv
import io
from datetime import date, datetime
from decimal import Decimal
import re
import unicodedata
from typing import Annotated
from uuid import UUID
from xml.etree.ElementTree import ParseError
from zipfile import BadZipFile

import psycopg
from fastapi import APIRouter, Depends, FastAPI, File, HTTPException, Query, Response, UploadFile, status
from fastapi.middleware.cors import CORSMiddleware
from psycopg.errors import CheckViolation, UniqueViolation
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from .config import settings
from .db import get_connection
from .domain.calendar import month_bounds, next_month
from .domain.commitments import (
    commitment_due_day,
    commitment_due_month,
    first_commitment_occurrence,
    next_commitment_due_date,
    next_projected_commitment_date,
    projected_commitment_date,
)
from .domain.budgets import calculate_allocation
from .schemas import (
    BudgetAllocationMode,
    BudgetAllocationRead,
    BudgetAllocationUpdate,
    BudgetBaseMode,
    BudgetDashboardRead,
    BudgetSettingsRead,
    BudgetSettingsUpdate,
    BudgetSummaryRead,
    CategoryCreate,
    CategoryRead,
    CategoryUpdate,
    CommitmentCreate,
    CommitmentRead,
    CommitmentRecordCreate,
    CommitmentRecordResult,
    CommitmentPreview,
    CommitmentUpdate,
    DashboardPeriod,
    DashboardRead,
    Direction,
    TransactionCreate,
    TransactionRead,
    TransactionUpdate,
    UserSettingsRead,
    UserSettingsUpdate,
)
from .security import current_user_id


app = FastAPI(
    title="Cifro API",
    version="0.1.0",
    description="API pessoal do gerenciador financeiro Cifro.",
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=settings.allowed_origins,
    allow_credentials=False,
    allow_methods=["GET", "POST", "PATCH", "DELETE", "OPTIONS"],
    allow_headers=["Content-Type", "Authorization"],
)

router = APIRouter(
    prefix="/api/v1",
)


def as_money(value: Decimal | None) -> Decimal:
    return value or Decimal("0.00")


IMPORT_MAX_BYTES = 10 * 1024 * 1024
IMPORT_MAX_SHEETS = 24
IMPORT_MAX_ROWS_PER_SHEET = 10_000
IMPORT_MAX_COLUMNS_PER_SHEET = 100
IMPORT_PREVIEW_LIMIT = 100
IMPORT_FIELDS = ("date", "description", "amount", "direction", "income", "expense", "category", "status", "notes")
IMPORT_MONTH_NAMES = {
    "janeiro": 1,
    "fevereiro": 2,
    "marco": 3,
    "abril": 4,
    "maio": 5,
    "junho": 6,
    "julho": 7,
    "agosto": 8,
    "setembro": 9,
    "outubro": 10,
    "novembro": 11,
    "dezembro": 12,
    "jan": 1,
    "fev": 2,
    "mar": 3,
    "abr": 4,
    "mai": 5,
    "jun": 6,
    "jul": 7,
    "ago": 8,
    "set": 9,
    "out": 10,
    "nov": 11,
    "dez": 12,
}
IMPORT_ALIASES = {
    "date": ("data", "date", "dia", "quando", "ocorrido em", "ocorrencia"),
    "description": ("descricao", "historico", "detalhes", "lancamento", "movimento", "evento", "nome", "item", "item/servico", "servico", "produto"),
    "amount": ("valor", "amount", "total", "quantia", "preco", "preco total"),
    "direction": ("tipo", "type", "natureza", "operacao", "movimentacao", "movimento"),
    "income": ("entrada", "receita", "recebimento", "credito", "creditos"),
    "expense": ("saida", "despesa", "gasto", "pagamento", "debito", "debitos"),
    "category": ("categoria", "classificacao", "grupo", "category"),
    "status": ("status", "situacao", "estado"),
    "notes": ("observacoes", "observacao", "obs", "notas", "notes"),
}


class ImportLimitError(ValueError):
    """Raised when an import exceeds a deliberate resource limit."""


def safe_csv_text(value: object) -> str:
    """Keep user-controlled text from being interpreted as a spreadsheet formula."""
    text = "" if value is None else str(value)
    stripped = text.lstrip(" \t\r\n")
    if stripped[:1] in {"=", "+", "-", "@"} or text[:1] in {"\t", "\r", "\n"}:
        return "'" + text
    return text


def import_format(filename: str) -> str | None:
    lower_filename = filename.lower()
    if lower_filename.endswith(".csv"):
        return "csv"
    if lower_filename.endswith(".xlsx"):
        return "xlsx"
    return None


def normalize_import_text(value: object) -> str:
    text = "" if value is None else str(value).strip().lower()
    return "".join(
        character
        for character in unicodedata.normalize("NFD", text)
        if unicodedata.category(character) != "Mn"
    )


def import_cell_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return str(value).strip()


def find_import_header(rows: list[list[object]]) -> tuple[int | None, list[str], int]:
    best: tuple[int | None, list[str], int] = (None, [], 0)
    known_aliases = {alias for aliases in IMPORT_ALIASES.values() for alias in aliases}
    for index, row in enumerate(rows[:20]):
        headers = [import_cell_text(cell) for cell in row]
        normalized_headers = [normalize_import_text(header) for header in headers]
        score = sum(
            any(alias == header or alias in header for alias in known_aliases)
            for header in normalized_headers
            if header
        )
        if score > best[2]:
            best = (index, headers, score)
    if best[2] < 2:
        return None, [], best[2]
    return best


def detect_import_mapping(headers: list[str]) -> dict[str, str | None]:
    normalized_headers = [normalize_import_text(header) for header in headers]
    mapping: dict[str, str | None] = {}
    used: set[int] = set()
    for field in IMPORT_FIELDS:
        aliases = sorted(IMPORT_ALIASES[field], key=len, reverse=True)
        found_index = None
        for index, header in enumerate(normalized_headers):
            if index in used or not header:
                continue
            if any(alias == header or alias in header for alias in aliases):
                found_index = index
                break
        mapping[field] = headers[found_index] if found_index is not None else None
        if found_index is not None:
            used.add(found_index)
    return mapping


def count_import_month_headers(rows: list[list[object]]) -> int:
    found_months: set[int] = set()
    for row in rows[:8]:
        for value in row:
            normalized = normalize_import_text(value)
            if normalized in IMPORT_MONTH_NAMES:
                found_months.add(IMPORT_MONTH_NAMES[normalized])
    return len(found_months)


def count_nonempty_import_rows(rows: list[list[object]], start: int = 0) -> int:
    return sum(1 for values in rows[start:] if any(import_cell_text(value) for value in values))


def classify_import_sheet(
    rows: list[list[object]],
    header_index: int | None,
    headers: list[str],
) -> tuple[str, str, int]:
    mapping = detect_import_mapping(headers) if header_index is not None else {}
    indexes = import_mapping_indexes(headers, mapping) if header_index is not None else {}
    has_transaction_shape = (
        header_index is not None
        and indexes.get("date") is not None
        and indexes.get("description") is not None
        and (
            indexes.get("amount") is not None
            or indexes.get("income") is not None
            or indexes.get("expense") is not None
        )
    )
    month_headers = count_import_month_headers(rows)
    if has_transaction_shape:
        return "transactions", "Encontramos linhas com data, descrição e valor.", 90
    if month_headers >= 3:
        return "summary", "Encontramos meses nas colunas; este parece ser um resumo mensal ou orçamento.", 95
    if header_index is not None:
        return "unknown", "Há um cabeçalho, mas faltam campos suficientes para identificar movimentações.", 45
    return "unknown", "Não encontrei uma estrutura de movimentações ou de resumo mensal reconhecível.", 20


def import_mapping_indexes(headers: list[str], mapping: dict[str, str | None]) -> dict[str, int | None]:
    return {
        field: headers.index(column) if column in headers else None
        for field, column in mapping.items()
    }


def parse_import_date(value: object) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = import_cell_text(value)
    if not text:
        return None
    for pattern in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, pattern).date().isoformat()
        except ValueError:
            continue
    return None


IMPORT_MONTHS = {
    "janeiro": 1,
    "fevereiro": 2,
    "marco": 3,
    "abril": 4,
    "maio": 5,
    "junho": 6,
    "julho": 7,
    "agosto": 8,
    "setembro": 9,
    "outubro": 10,
    "novembro": 11,
    "dezembro": 12,
}


def parse_import_sheet_period(sheet_name: str) -> tuple[int, int] | None:
    normalized = normalize_import_text(sheet_name)
    month_pattern = "|".join(IMPORT_MONTHS)
    month_match = re.search(rf"({month_pattern})\D*(20\d{{2}})", normalized)
    if month_match:
        return int(month_match.group(2)), IMPORT_MONTHS[month_match.group(1)]

    numeric_match = re.search(r"(20\d{2})\D{1,3}(0?[1-9]|1[0-2])", normalized)
    if numeric_match:
        return int(numeric_match.group(1)), int(numeric_match.group(2))
    return None


def parse_import_partial_date(value: object, sheet_name: str) -> str | None:
    text = import_cell_text(value)
    match = re.fullmatch(r"(\d{1,2})[/-](\d{1,2})", text)
    period = parse_import_sheet_period(sheet_name)
    if not match or not period:
        return None

    first, second = (int(match.group(1)), int(match.group(2)))
    year, sheet_month = period
    if second == sheet_month and first <= 31:
        day = first
    elif first == sheet_month and second <= 31:
        day = second
    else:
        return None
    try:
        return date(year, sheet_month, day).isoformat()
    except ValueError:
        return None


def parse_import_amount(value: object) -> Decimal | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float, Decimal)):
        try:
            return Decimal(str(value)).quantize(Decimal("0.01"))
        except (ArithmeticError, ValueError):
            return None

    text = import_cell_text(value).replace("R$", "").replace("r$", "").replace(" ", "")
    if not text:
        return None
    negative = text.startswith("-") or (text.startswith("(") and text.endswith(")"))
    text = text.strip("()")
    text = re.sub(r"[^0-9,.]", "", text)
    if not text:
        return None
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")
    elif text.count(".") > 1:
        text = text.replace(".", "")
    try:
        amount = Decimal(text).quantize(Decimal("0.01"))
    except (ArithmeticError, ValueError):
        return None
    return -amount if negative else amount


def parse_import_direction(value: object) -> str | None:
    normalized = normalize_import_text(value)
    if normalized in {"income", "recebimento", "recebimentos", "entrada", "receita", "credito", "creditos", "+"}:
        return "income"
    if normalized in {"expense", "gasto", "gastos", "saida", "despesa", "pagamento", "debito", "debitos", "-"}:
        return "expense"
    return None


def parse_import_status(value: object) -> str:
    normalized = normalize_import_text(value)
    if normalized in {"planned", "pendente", "pendentes", "previsto", "prevista"}:
        return "planned"
    return "completed"


def normalize_import_row(
    values: list[object],
    indexes: dict[str, int | None],
    sheet_name: str,
    row_number: int,
) -> dict:
    def value_for(field: str) -> object:
        index = indexes.get(field)
        return values[index] if index is not None and index < len(values) else None

    errors: list[str] = []
    warnings: list[str] = []
    raw_description = import_cell_text(value_for("description"))
    parsed_date = parse_import_date(value_for("date"))
    if not parsed_date:
        partial_date = parse_import_partial_date(value_for("date"), sheet_name)
        if partial_date:
            parsed_date = partial_date
            warnings.append("Ano/mês complementados pelo nome da aba")
    direction = parse_import_direction(value_for("direction"))
    income = parse_import_amount(value_for("income"))
    expense = parse_import_amount(value_for("expense"))
    amount = parse_import_amount(value_for("amount"))

    if not raw_description:
        errors.append("Descrição ausente")
    if not parsed_date:
        errors.append("Data ausente ou inválida")

    has_income = income is not None and income != 0
    has_expense = expense is not None and expense != 0
    if has_income and has_expense:
        errors.append("Entrada e saída preenchidas simultaneamente")
        amount = None
        direction = None
    elif has_income:
        amount = abs(income)
        direction = "income"
    elif has_expense:
        amount = abs(expense)
        direction = "expense"
    elif amount is not None and amount < 0:
        amount = abs(amount)
        direction = direction or "expense"

    if amount is None or amount <= 0:
        errors.append("Valor ausente ou inválido")
    if direction is None:
        errors.append("Tipo não identificado")
    if indexes.get("status") is None:
        warnings.append("Status não encontrado; será tratado como concluído")

    return {
        "sheet_name": sheet_name,
        "source_row": row_number,
        "date": parsed_date,
        "description": raw_description,
        "amount": f"{amount:.2f}" if amount is not None else None,
        "direction": direction,
        "category": import_cell_text(value_for("category")) or None,
        "status": parse_import_status(value_for("status")),
        "notes": import_cell_text(value_for("notes")) or None,
        "errors": errors,
        "warnings": warnings,
        "valid": not errors,
    }


def read_import_sheets(filename: str, content: bytes) -> list[tuple[str, list[list[object]]]]:
    file_type = import_format(filename)
    if file_type == "csv":
        try:
            text = content.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = content.decode("cp1252")
        sample = text[:4096]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=";,\t,")
        except csv.Error:
            dialect = csv.excel
            dialect.delimiter = ";"
        rows = []
        for row_number, row in enumerate(csv.reader(io.StringIO(text), dialect), start=1):
            if row_number > IMPORT_MAX_ROWS_PER_SHEET:
                raise ImportLimitError("A planilha excede o limite de linhas permitido")
            if len(row) > IMPORT_MAX_COLUMNS_PER_SHEET:
                raise ImportLimitError("A planilha excede o limite de colunas permitido")
            rows.append(row)
        return [(filename.rsplit(".", 1)[0], rows)]

    if file_type == "xlsx":
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        if len(workbook.worksheets) > IMPORT_MAX_SHEETS:
            workbook.close()
            raise ImportLimitError("A planilha excede o limite de abas permitido")
        sheets = []
        for worksheet in workbook.worksheets:
            rows = []
            for row_number, row in enumerate(
                worksheet.iter_rows(
                    values_only=True,
                    max_row=IMPORT_MAX_ROWS_PER_SHEET + 1,
                    max_col=IMPORT_MAX_COLUMNS_PER_SHEET + 1,
                ),
                start=1,
            ):
                if row_number > IMPORT_MAX_ROWS_PER_SHEET:
                    if any(import_cell_text(value) for value in row):
                        workbook.close()
                        raise ImportLimitError("A planilha excede o limite de linhas permitido")
                    break
                if len(row) > IMPORT_MAX_COLUMNS_PER_SHEET and any(row[IMPORT_MAX_COLUMNS_PER_SHEET:]):
                    workbook.close()
                    raise ImportLimitError("A planilha excede o limite de colunas permitido")
                rows.append(list(row[:IMPORT_MAX_COLUMNS_PER_SHEET]))
            sheets.append((worksheet.title, rows))
        workbook.close()
        return sheets

    raise HTTPException(status_code=400, detail="Use um arquivo CSV ou XLSX")


COMMITMENT_COLUMNS = """
  c.id, c.user_id, c.category_id, c.name, c.commitment_type, c.direction,
  c.amount, c.frequency, c.due_rule, c.due_day, c.due_month, c.business_day_number, c.starts_on,
  c.next_due_on, c.ends_on, c.total_installments, c.current_installment,
  c.is_active, c.created_at
"""


def process_due_commitments(connection, today: date | None = None) -> int:
    """Create today's due occurrences once and advance their schedules.

    This intentionally does not backfill every missed month. If the service was
    offline, the next run creates the current occurrence instead of inventing a
    large historical block of salary or expenses.
    """
    today = today or date.today()
    rows = connection.execute(
        f"""
        select {COMMITMENT_COLUMNS},
          coalesce(us.auto_confirm_income, false) as auto_confirm_income
        from public.commitments c
        left join public.user_settings us on us.user_id = c.user_id
        where c.is_active = true
        order by c.user_id, c.next_due_on, c.name
        for update of c
        """
    ).fetchall()
    processed = 0

    for commitment in rows:
        if commitment["commitment_type"] == "installment":
            occurrence_on = commitment["next_due_on"]
        else:
            occurrence_on = projected_commitment_date(commitment, today.year, today.month)

        if occurrence_on is None or occurrence_on > today:
            continue

        existing = connection.execute(
            """
            select id
            from public.transactions
            where commitment_id = %s and user_id = %s and occurred_on = %s
            limit 1
            """,
            (commitment["id"], commitment["user_id"], occurrence_on),
        ).fetchone()
        if existing:
            continue

        transaction_status = (
            "completed"
            if commitment["direction"] == Direction.INCOME and commitment["auto_confirm_income"]
            else "planned"
        )
        connection.execute(
            """
            insert into public.transactions (
              user_id, category_id, commitment_id, description, amount,
              direction, occurred_on, status, notes
            )
            values (%s, %s, %s, %s, %s, %s, %s, %s, %s)
            """,
            (
                commitment["user_id"],
                commitment["category_id"],
                commitment["id"],
                commitment["name"],
                commitment["amount"],
                commitment["direction"],
                occurrence_on,
                transaction_status,
                "Gerado a partir do planejamento",
            ),
        )

        commitment_for_advance = dict(commitment)
        commitment_for_advance["next_due_on"] = occurrence_on
        next_due_on = next_commitment_due_date(commitment_for_advance)
        is_installment = commitment["commitment_type"] == "installment"
        current_installment = commitment["current_installment"]
        if is_installment and current_installment >= commitment["total_installments"]:
            is_active = False
        else:
            is_active = next_due_on is not None and not (
                commitment["ends_on"] and next_due_on > commitment["ends_on"]
            )
        if is_installment and is_active:
            current_installment += 1

        connection.execute(
            """
            update public.commitments
            set next_due_on = %s, current_installment = %s, is_active = %s
            where id = %s and user_id = %s
            """,
            (
                next_due_on or commitment["next_due_on"],
                current_installment,
                is_active,
                commitment["id"],
                commitment["user_id"],
            ),
        )
        processed += 1

    return processed


@app.get("/health")
def health() -> dict[str, str]:
    return {"status": "ok", "service": "cifro-api"}


@router.get("/settings", response_model=UserSettingsRead)
def get_user_settings(user_id: UUID = Depends(current_user_id)) -> dict:
    with get_connection() as connection:
        row = connection.execute(
            """
            select auto_confirm_income, default_due_rule,
              default_business_day_number, updated_at
            from public.user_settings
            where user_id = %s
            """,
            (user_id,),
        ).fetchone()
        if not row:
            row = connection.execute(
                """
                insert into public.user_settings (user_id)
                values (%s)
                returning auto_confirm_income, default_due_rule,
                  default_business_day_number, updated_at
                """,
                (user_id,),
            ).fetchone()
    return row


@router.patch("/settings", response_model=UserSettingsRead)
def update_user_settings(
    payload: UserSettingsUpdate,
    user_id: UUID = Depends(current_user_id),
) -> dict:
    with get_connection() as connection:
        row = connection.execute(
            """
            insert into public.user_settings (
              user_id, auto_confirm_income, default_due_rule,
              default_business_day_number, updated_at
            )
            values (%s, %s, %s, %s, now())
            on conflict (user_id) do update set
              auto_confirm_income = excluded.auto_confirm_income,
              default_due_rule = excluded.default_due_rule,
              default_business_day_number = excluded.default_business_day_number,
              updated_at = now()
            returning auto_confirm_income, default_due_rule,
              default_business_day_number, updated_at
            """,
            (
                user_id,
                payload.auto_confirm_income,
                payload.default_due_rule.value,
                payload.default_business_day_number,
            ),
        ).fetchone()
    return row


@router.get("/categories", response_model=list[CategoryRead])
def list_categories(user_id: UUID = Depends(current_user_id)) -> list[dict]:
    with get_connection() as connection:
        rows = connection.execute(
            """
            select id, name, kind, is_active, created_at
            from public.categories
            where user_id = %s and is_active = true
            order by name asc
            """,
            (user_id,),
        ).fetchall()
    return list(rows)


@router.post("/categories", response_model=CategoryRead, status_code=status.HTTP_201_CREATED)
def create_category(payload: CategoryCreate, user_id: UUID = Depends(current_user_id)) -> dict:
    try:
        with get_connection() as connection:
            return connection.execute(
                """
                insert into public.categories (user_id, name, kind)
                values (%s, %s, %s)
                returning id, name, kind, is_active, created_at
                """,
                (user_id, payload.name.strip(), payload.kind.value),
            ).fetchone()
    except UniqueViolation as error:
        raise HTTPException(status_code=409, detail="Category already exists") from error


@router.patch("/categories/{category_id}", response_model=CategoryRead)
def update_category(
    category_id: UUID,
    payload: CategoryUpdate,
    user_id: UUID = Depends(current_user_id),
) -> dict:
    try:
        with get_connection() as connection:
            row = connection.execute(
                """
                update public.categories
                set name = %s, kind = %s
                where id = %s and user_id = %s and is_active = true
                returning id, name, kind, is_active, created_at
                """,
                (payload.name.strip(), payload.kind.value, category_id, user_id),
            ).fetchone()
    except UniqueViolation as error:
        raise HTTPException(status_code=409, detail="Category already exists") from error
    except CheckViolation as error:
        raise HTTPException(
            status_code=409,
            detail="Category is being used by the budget distribution",
        ) from error

    if not row:
        raise HTTPException(status_code=404, detail="Category not found")
    return row


@router.delete("/categories/{category_id}", status_code=status.HTTP_204_NO_CONTENT)
def archive_category(category_id: UUID, user_id: UUID = Depends(current_user_id)) -> None:
    with get_connection() as connection:
        row = connection.execute(
            """
            update public.categories
            set is_active = false
            where id = %s and user_id = %s and is_active = true
            returning id
            """,
            (category_id, user_id),
        ).fetchone()

    if not row:
        raise HTTPException(status_code=404, detail="Category not found")


def get_budget_settings(connection, user_id: UUID) -> dict:
    connection.execute(
        """
        insert into public.budget_settings (user_id)
        values (%s)
        on conflict (user_id) do nothing
        """,
        (user_id,),
    )
    return connection.execute(
        """
        select
          bs.base_mode, bs.income_category_id, bs.manual_amount,
          bs.updated_at, c.name as income_category_name
        from public.budget_settings bs
        left join public.categories c
          on c.id = bs.income_category_id and c.user_id = bs.user_id
        where bs.user_id = %s
        """,
        (user_id,),
    ).fetchone()


def get_budget_month(
    connection,
    user_id: UUID,
    year: int,
    month: int,
) -> dict:
    month_date = date(year, month, 1)
    get_budget_settings(connection, user_id)
    connection.execute(
        "select pg_catalog.pg_advisory_xact_lock(pg_catalog.hashtextextended(%s::text, 0))",
        (f"budget-month:{user_id}:{month_date.isoformat()}",),
    )
    row = connection.execute(
        """
        select
          bm.id, bm.base_mode, bm.income_category_id, bm.manual_amount,
          bm.created_at, bm.updated_at, c.name as income_category_name
        from public.budget_months bm
        left join public.categories c
          on c.id = bm.income_category_id and c.user_id = bm.user_id
        where bm.user_id = %s and bm.month = %s
        """,
        (user_id, month_date),
    ).fetchone()
    if row:
        return row

    row = connection.execute(
        """
        insert into public.budget_months (
          user_id, month, base_mode, income_category_id, manual_amount
        )
        select user_id, %s, base_mode, income_category_id, manual_amount
        from public.budget_settings
        where user_id = %s
        returning id
        """,
        (month_date, user_id),
    ).fetchone()
    connection.execute(
        """
        insert into public.budget_month_allocations (
          budget_month_id, user_id, category_id,
          allocation_mode, percentage, fixed_amount
        )
        select %s, user_id, category_id,
          allocation_mode, percentage, fixed_amount
        from public.budget_allocations
        where user_id = %s
        """,
        (row["id"], user_id),
    )
    return connection.execute(
        """
        select
          bm.id, bm.base_mode, bm.income_category_id, bm.manual_amount,
          bm.created_at, bm.updated_at, c.name as income_category_name
        from public.budget_months bm
        left join public.categories c
          on c.id = bm.income_category_id and c.user_id = bm.user_id
        where bm.id = %s and bm.user_id = %s
        """,
        (row["id"], user_id),
    ).fetchone()


def budget_base_amount(
    connection,
    user_id: UUID,
    year: int,
    month: int,
    settings_row: dict,
) -> Decimal:
    if settings_row["base_mode"] == BudgetBaseMode.MANUAL:
        return as_money(settings_row["manual_amount"])

    period_start, period_end = month_bounds(year, month)
    if settings_row["base_mode"] == BudgetBaseMode.TOTAL_INCOME:
        row = connection.execute(
            """
            select coalesce(sum(amount), 0) as total
            from public.transactions
            where user_id = %s
              and status = 'completed'
              and direction = 'income'
              and occurred_on >= %s and occurred_on < %s
            """,
            (user_id, period_start, period_end),
        ).fetchone()
        return as_money(row["total"])

    commitment_rows = connection.execute(
        f"""
        select {COMMITMENT_COLUMNS}
        from public.commitments c
        where c.user_id = %s
          and c.category_id = %s
          and c.direction = 'income'
          and c.is_active = true
        """,
        (user_id, settings_row["income_category_id"]),
    ).fetchall()
    return as_money(sum(
        (
            row["amount"]
            for row in commitment_rows
            if projected_commitment_date(row, year, month) is not None
        ),
        Decimal("0.00"),
    ))


def build_budget_summary(connection, user_id: UUID, year: int, month: int) -> dict:
    period_start, period_end = month_bounds(year, month)
    settings_row = get_budget_month(connection, user_id, year, month)
    base_amount = budget_base_amount(connection, user_id, year, month, settings_row)

    allocation_rows = connection.execute(
        """
        select
          ba.category_id, c.name as category_name, ba.allocation_mode,
          ba.percentage, ba.fixed_amount,
          coalesce(sum(t.amount), 0) as actual_amount
        from public.budget_month_allocations ba
        join public.categories c
          on c.id = ba.category_id and c.user_id = ba.user_id
        left join public.transactions t
          on t.user_id = ba.user_id
         and t.category_id = ba.category_id
         and t.direction = 'expense'
         and t.status = 'completed'
         and t.occurred_on >= %s and t.occurred_on < %s
        where ba.user_id = %s and ba.budget_month_id = %s
        group by
          ba.category_id, c.name, ba.allocation_mode,
          ba.percentage, ba.fixed_amount
        order by c.name asc
        """,
        (period_start, period_end, user_id, settings_row["id"]),
    ).fetchall()

    allocations = []
    allocated_amount = Decimal("0.00")
    for row in allocation_rows:
        calculation = calculate_allocation(
            base_amount,
            row["allocation_mode"],
            row["percentage"],
            row["fixed_amount"],
            row["actual_amount"],
        )
        fixed_amount = calculation["fixed_amount"]
        percentage = calculation["percentage"]
        target_amount = calculation["target_amount"]
        actual_amount = calculation["actual_amount"]
        allocated_amount += target_amount
        allocations.append(
            BudgetAllocationRead(
                category_id=row["category_id"],
                category_name=row["category_name"],
                allocation_mode=row["allocation_mode"],
                percentage=percentage,
                fixed_amount=fixed_amount,
                target_amount=target_amount,
                actual_amount=actual_amount,
                remaining_amount=target_amount - actual_amount,
            )
        )

    allocations.sort(key=lambda allocation: (-allocation.target_amount, allocation.category_name.lower()))
    total_percentage = (
        allocated_amount * Decimal("100") / base_amount
    ).quantize(Decimal("0.01")) if base_amount > 0 else Decimal("0.00")
    unallocated_percentage = Decimal("100.00") - total_percentage
    unallocated_amount = base_amount - allocated_amount
    return {
        "month": f"{year:04d}-{month:02d}",
        "settings": BudgetSettingsRead(**settings_row),
        "base_amount": base_amount,
        "allocated_amount": allocated_amount,
        "total_percentage": total_percentage,
        "unallocated_percentage": unallocated_percentage,
        "unallocated_amount": unallocated_amount,
        "allocations": allocations,
    }


@router.get("/budget", response_model=BudgetSummaryRead)
def get_budget(
    year: Annotated[int, Query(ge=2000, le=2100)] | None = None,
    month: Annotated[int, Query(ge=1, le=12)] | None = None,
    user_id: UUID = Depends(current_user_id),
) -> dict:
    today = date.today()
    selected_year = year or today.year
    selected_month = month or today.month
    with get_connection() as connection:
        return build_budget_summary(connection, user_id, selected_year, selected_month)


@router.patch("/budget/settings", response_model=BudgetSettingsRead)
def update_budget_settings(
    payload: BudgetSettingsUpdate,
    year: Annotated[int, Query(ge=2000, le=2100)] | None = None,
    month: Annotated[int, Query(ge=1, le=12)] | None = None,
    user_id: UUID = Depends(current_user_id),
) -> dict:
    today = date.today()
    selected_year = year or today.year
    selected_month = month or today.month
    try:
        with get_connection() as connection:
            if payload.base_mode == BudgetBaseMode.CATEGORY_INCOME:
                category = connection.execute(
                    """
                    select id
                    from public.categories
                    where id = %s and user_id = %s and is_active = true
                      and kind in ('income', 'both')
                    """,
                    (payload.income_category_id, user_id),
                ).fetchone()
                if not category:
                    raise HTTPException(status_code=400, detail="Income category not found")

            budget_month = get_budget_month(
                connection,
                user_id,
                selected_year,
                selected_month,
            )
            connection.execute(
                """
                update public.budget_months
                set base_mode = %s,
                    income_category_id = %s,
                    manual_amount = %s,
                    updated_at = now()
                where id = %s and user_id = %s
                """,
                (
                    payload.base_mode.value,
                    payload.income_category_id,
                    payload.manual_amount,
                    budget_month["id"],
                    user_id,
                ),
            )
            return get_budget_month(connection, user_id, selected_year, selected_month)
    except CheckViolation as error:
        raise HTTPException(status_code=400, detail="Invalid budget base") from error


@router.post("/budget/template", status_code=status.HTTP_204_NO_CONTENT)
def update_budget_template(
    year: Annotated[int, Query(ge=2000, le=2100)] | None = None,
    month: Annotated[int, Query(ge=1, le=12)] | None = None,
    user_id: UUID = Depends(current_user_id),
) -> Response:
    today = date.today()
    selected_year = year or today.year
    selected_month = month or today.month
    with get_connection() as connection:
        connection.execute(
            "select pg_catalog.pg_advisory_xact_lock(pg_catalog.hashtextextended(%s::text, 0))",
            (user_id,),
        )
        budget_month = get_budget_month(
            connection,
            user_id,
            selected_year,
            selected_month,
        )
        connection.execute(
            """
            insert into public.budget_settings (
              user_id, base_mode, income_category_id, manual_amount, updated_at
            )
            values (%s, %s, %s, %s, now())
            on conflict (user_id) do update set
              base_mode = excluded.base_mode,
              income_category_id = excluded.income_category_id,
              manual_amount = excluded.manual_amount,
              updated_at = now()
            """,
            (
                user_id,
                budget_month["base_mode"],
                budget_month["income_category_id"],
                budget_month["manual_amount"],
            ),
        )
        connection.execute(
            "delete from public.budget_allocations where user_id = %s",
            (user_id,),
        )
        connection.execute(
            """
            insert into public.budget_allocations (
              user_id, category_id, allocation_mode, percentage, fixed_amount
            )
            select user_id, category_id, allocation_mode, percentage, fixed_amount
            from public.budget_month_allocations
            where budget_month_id = %s and user_id = %s
            """,
            (budget_month["id"], user_id),
        )
    return Response(status_code=status.HTTP_204_NO_CONTENT)


@router.patch("/budget/allocations/{category_id}", status_code=status.HTTP_204_NO_CONTENT)
def update_budget_allocation(
    category_id: UUID,
    payload: BudgetAllocationUpdate,
    year: Annotated[int, Query(ge=2000, le=2100)] | None = None,
    month: Annotated[int, Query(ge=1, le=12)] | None = None,
    user_id: UUID = Depends(current_user_id),
) -> Response:
    today = date.today()
    selected_year = year or today.year
    selected_month = month or today.month
    try:
        with get_connection() as connection:
            category = connection.execute(
                """
                select id
                from public.categories
                where id = %s and user_id = %s and is_active = true
                  and kind in ('expense', 'both')
                """,
                (category_id, user_id),
            ).fetchone()
            if not category:
                raise HTTPException(status_code=400, detail="Expense category not found")

            connection.execute(
                "select pg_catalog.pg_advisory_xact_lock(pg_catalog.hashtextextended(%s::text, 0))",
                (user_id,),
            )
            budget_month = get_budget_month(
                connection,
                user_id,
                selected_year,
                selected_month,
            )
            connection.execute(
                """
                insert into public.budget_month_allocations (
                  budget_month_id, user_id, category_id,
                  allocation_mode, percentage, fixed_amount
                )
                values (%s, %s, %s, %s, %s, %s)
                on conflict (budget_month_id, category_id) do update set
                  allocation_mode = excluded.allocation_mode,
                  percentage = excluded.percentage,
                  fixed_amount = excluded.fixed_amount,
                  updated_at = now()
                """,
                (
                    budget_month["id"],
                    user_id,
                    category_id,
                    payload.allocation_mode.value,
                    payload.percentage,
                    payload.fixed_amount,
                ),
            )
    except CheckViolation as error:
        raise HTTPException(status_code=400, detail="Invalid budget allocation") from error
    return Response(status_code=status.HTTP_204_NO_CONTENT)


@router.delete("/budget/allocations/{category_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_budget_allocation(
    category_id: UUID,
    year: Annotated[int, Query(ge=2000, le=2100)] | None = None,
    month: Annotated[int, Query(ge=1, le=12)] | None = None,
    user_id: UUID = Depends(current_user_id),
) -> Response:
    today = date.today()
    selected_year = year or today.year
    selected_month = month or today.month
    with get_connection() as connection:
        budget_month = get_budget_month(
            connection,
            user_id,
            selected_year,
            selected_month,
        )
        connection.execute(
            """
            delete from public.budget_month_allocations
            where budget_month_id = %s and user_id = %s and category_id = %s
            """,
            (budget_month["id"], user_id, category_id),
        )
    return Response(status_code=status.HTTP_204_NO_CONTENT)


def validate_category_for_direction(
    connection,
    category_id: UUID | None,
    direction: str,
    user_id: UUID,
) -> dict | None:
    if not category_id:
        return None

    category = connection.execute(
        """
        select id, name, kind, is_active
        from public.categories
        where id = %s and user_id = %s
        """,
        (category_id, user_id),
    ).fetchone()
    if not category or not category["is_active"]:
        raise HTTPException(status_code=400, detail="Category not found or inactive")
    if category["kind"] not in (direction, "both"):
        raise HTTPException(
            status_code=400,
            detail="Category is not compatible with this direction",
        )
    return category


def validate_commitment_for_transaction(
    connection,
    commitment_id: UUID | None,
    direction: str,
    category_id: UUID | None,
    user_id: UUID,
) -> dict | None:
    if not commitment_id:
        return None

    commitment = connection.execute(
        """
        select id, category_id, direction, is_active
        from public.commitments
        where id = %s and user_id = %s
        """,
        (commitment_id, user_id),
    ).fetchone()
    if not commitment or not commitment["is_active"]:
        raise HTTPException(status_code=400, detail="Commitment not found or inactive")
    if commitment["direction"] != direction:
        raise HTTPException(
            status_code=400,
            detail="Commitment is not compatible with this direction",
        )
    if category_id and commitment["category_id"] and category_id != commitment["category_id"]:
        raise HTTPException(
            status_code=400,
            detail="Transaction category must match its commitment category",
        )
    return commitment


@router.get("/commitments", response_model=list[CommitmentRead])
def list_commitments(user_id: UUID = Depends(current_user_id)) -> list[dict]:
    with get_connection() as connection:
        rows = connection.execute(
            """
            select
              c.id, c.name, c.commitment_type, c.direction, c.amount,
              c.frequency, c.due_rule, c.due_day, c.due_month, c.business_day_number,
              c.starts_on, c.next_due_on, c.ends_on,
              c.category_id, c.total_installments, c.current_installment,
              c.is_active, c.created_at, cat.name as category_name
            from public.commitments c
            left join public.categories cat
              on cat.id = c.category_id and cat.user_id = c.user_id
            where c.user_id = %s and c.is_active = true
            order by c.next_due_on asc, c.name asc
            """,
            (user_id,),
        ).fetchall()
    commitments = []
    for row in rows:
        projected_date = next_projected_commitment_date(row, date.today())
        if projected_date is None and row["commitment_type"] == "installment":
            projected_date = row["next_due_on"]
        if projected_date is None:
            continue
        projected = dict(row)
        projected["next_due_on"] = projected_date
        commitments.append(projected)
    commitments.sort(key=lambda row: (row["next_due_on"], row["name"].lower()))
    return commitments


@router.post("/commitments", response_model=CommitmentRead, status_code=status.HTTP_201_CREATED)
def create_commitment(payload: CommitmentCreate, user_id: UUID = Depends(current_user_id)) -> dict:
    next_due_on = first_commitment_occurrence(
        payload.starts_on,
        payload.frequency.value,
        payload.due_rule.value,
        payload.due_day,
        payload.due_month,
        payload.business_day_number,
    )
    if next_due_on is None:
        raise HTTPException(status_code=400, detail="Could not calculate the commitment occurrence")

    with get_connection() as connection:
        category = validate_category_for_direction(
            connection,
            payload.category_id,
            payload.direction.value,
            user_id,
        )
        row = connection.execute(
            """
            insert into public.commitments (
              user_id, category_id, name, commitment_type, direction, amount,
              frequency, due_rule, due_day, due_month, business_day_number,
              starts_on, next_due_on, ends_on,
              total_installments, current_installment
            )
            values (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            returning id, name, commitment_type, direction, amount, frequency,
              due_rule, due_day, due_month, business_day_number, starts_on, next_due_on, ends_on, category_id,
              total_installments, current_installment, is_active, created_at
            """,
            (
                user_id,
                payload.category_id,
                payload.name.strip(),
                payload.commitment_type.value,
                payload.direction.value,
                payload.amount,
                payload.frequency.value,
                payload.due_rule.value,
                payload.due_day,
                payload.due_month,
                payload.business_day_number,
                payload.starts_on,
                next_due_on,
                payload.ends_on,
                payload.total_installments,
                payload.current_installment,
            ),
        ).fetchone()
    row["category_name"] = category["name"] if category else None
    return row


@router.patch("/commitments/{commitment_id}", response_model=CommitmentRead)
def update_commitment(
    commitment_id: UUID,
    payload: CommitmentUpdate,
    user_id: UUID = Depends(current_user_id),
) -> dict:
    next_due_on = first_commitment_occurrence(
        payload.starts_on,
        payload.frequency.value,
        payload.due_rule.value,
        payload.due_day,
        payload.due_month,
        payload.business_day_number,
    )
    if next_due_on is None:
        raise HTTPException(status_code=400, detail="Could not calculate the commitment occurrence")

    with get_connection() as connection:
        category = validate_category_for_direction(
            connection,
            payload.category_id,
            payload.direction.value,
            user_id,
        )
        row = connection.execute(
            """
            update public.commitments
            set category_id = %s, name = %s, commitment_type = %s,
                direction = %s, amount = %s, frequency = %s,
                due_rule = %s, due_day = %s, due_month = %s, business_day_number = %s,
                starts_on = %s, next_due_on = %s, ends_on = %s,
                total_installments = %s, current_installment = %s
            where id = %s and user_id = %s and is_active = true
            returning id, name, commitment_type, direction, amount, frequency,
              due_rule, due_day, due_month, business_day_number, starts_on, next_due_on, ends_on, category_id,
              total_installments, current_installment, is_active, created_at
            """,
            (
                payload.category_id,
                payload.name.strip(),
                payload.commitment_type.value,
                payload.direction.value,
                payload.amount,
                payload.frequency.value,
                payload.due_rule.value,
                payload.due_day,
                payload.due_month,
                payload.business_day_number,
                payload.starts_on,
                next_due_on,
                payload.ends_on,
                payload.total_installments,
                payload.current_installment,
                commitment_id,
                user_id,
            ),
        ).fetchone()

    if not row:
        raise HTTPException(status_code=404, detail="Commitment not found")
    row["category_name"] = category["name"] if category else None
    return row


@router.post("/commitments/{commitment_id}/record", response_model=CommitmentRecordResult)
def record_commitment(
    commitment_id: UUID,
    payload: CommitmentRecordCreate,
    user_id: UUID = Depends(current_user_id),
) -> dict:
    with get_connection() as connection:
        commitment = connection.execute(
            """
            select
              id, user_id, category_id, name, commitment_type, direction, amount,
              frequency, due_rule, due_day, due_month, business_day_number, starts_on, next_due_on,
              ends_on, total_installments, current_installment, is_active, created_at
            from public.commitments
            where id = %s and user_id = %s and is_active = true
            for update
            """,
            (commitment_id, user_id),
        ).fetchone()

        if not commitment:
            raise HTTPException(status_code=404, detail="Commitment not found")

        pending_transaction = connection.execute(
            """
            select id, description, amount, direction, occurred_on,
              category_id, commitment_id, status, notes, created_at, updated_at
            from public.transactions
            where commitment_id = %s and user_id = %s and status = 'planned'
            order by occurred_on asc, created_at asc
            limit 1
            """,
            (commitment_id, user_id),
        ).fetchone()
        if pending_transaction:
            confirmed_transaction = connection.execute(
                """
                update public.transactions
                set status = 'completed'
                where id = %s and user_id = %s
                returning id, description, amount, direction, occurred_on,
                  category_id, commitment_id, status, notes, created_at, updated_at
                """,
                (pending_transaction["id"], user_id),
            ).fetchone()
            category_name = None
            if commitment["category_id"]:
                category = connection.execute(
                    """
                    select name
                    from public.categories
                    where id = %s and user_id = %s
                    """,
                    (commitment["category_id"], user_id),
                ).fetchone()
                category_name = category["name"] if category else None

            updated_commitment = dict(commitment)
            updated_commitment["category_name"] = category_name
            confirmed_transaction["category_name"] = category_name
            return {"transaction": confirmed_transaction, "commitment": updated_commitment}

        scheduled_due_on = next_projected_commitment_date(commitment, date.today())
        if scheduled_due_on is None or commitment["commitment_type"] == "installment":
            scheduled_due_on = commitment["next_due_on"]
        occurrence_on = payload.occurred_on or scheduled_due_on
        duplicate = connection.execute(
            """
            select id
            from public.transactions
            where commitment_id = %s and user_id = %s
              and occurred_on = %s and status = 'completed'
            limit 1
            """,
            (commitment_id, user_id, occurrence_on),
        ).fetchone()
        if duplicate:
            raise HTTPException(status_code=409, detail="This commitment occurrence is already recorded")

        transaction = connection.execute(
            """
            insert into public.transactions (
              user_id, category_id, commitment_id, description, amount,
              direction, occurred_on, status, notes
            )
            values (%s, %s, %s, %s, %s, %s, %s, 'completed', %s)
            returning id, description, amount, direction, occurred_on,
              category_id, commitment_id, status, notes, created_at, updated_at
            """,
            (
                user_id,
                commitment["category_id"],
                commitment_id,
                commitment["name"],
                commitment["amount"],
                commitment["direction"],
                occurrence_on,
                "Registrado a partir do planejamento",
            ),
        ).fetchone()

        commitment_for_advance = dict(commitment)
        commitment_for_advance["next_due_on"] = scheduled_due_on
        next_due_on = next_commitment_due_date(commitment_for_advance)
        is_installment = commitment["commitment_type"] == "installment"
        current_installment = commitment["current_installment"]
        if is_installment and current_installment >= commitment["total_installments"]:
            is_active = False
        else:
            is_active = next_due_on is not None and not (
                commitment["ends_on"] and next_due_on > commitment["ends_on"]
            )

        if is_installment and is_active:
            current_installment += 1

        updated_commitment = connection.execute(
            """
            update public.commitments
            set next_due_on = %s, current_installment = %s, is_active = %s
            where id = %s and user_id = %s
            returning id, name, commitment_type, direction, amount, frequency,
              due_rule, due_day, due_month, business_day_number, starts_on, next_due_on, ends_on,
              category_id, total_installments, current_installment,
              is_active, created_at
            """,
            (
                next_due_on or commitment["next_due_on"],
                current_installment,
                is_active,
                commitment_id,
                user_id,
            ),
        ).fetchone()

        category_name = None
        if commitment["category_id"]:
            category = connection.execute(
                """
                select name
                from public.categories
                where id = %s and user_id = %s
                """,
                (commitment["category_id"], user_id),
            ).fetchone()
            category_name = category["name"] if category else None

    transaction["category_name"] = category_name
    updated_commitment["category_name"] = category_name
    return {"transaction": transaction, "commitment": updated_commitment}


@router.delete("/commitments/{commitment_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_commitment(commitment_id: UUID, user_id: UUID = Depends(current_user_id)) -> None:
    with get_connection() as connection:
        row = connection.execute(
            """
            delete from public.commitments
            where id = %s and user_id = %s
            returning id
            """,
            (commitment_id, user_id),
        ).fetchone()

    if not row:
        raise HTTPException(status_code=404, detail="Commitment not found")


@router.get("/transactions", response_model=list[TransactionRead])
def list_transactions(
    limit: Annotated[int, Query(ge=1, le=100)] = 20,
    offset: Annotated[int, Query(ge=0)] = 0,
    user_id: UUID = Depends(current_user_id),
) -> list[dict]:
    with get_connection() as connection:
        rows = connection.execute(
            """
            select
              t.id, t.description, t.amount, t.direction, t.occurred_on,
              t.category_id, t.commitment_id, t.status, t.notes,
              t.created_at, t.updated_at, c.name as category_name
            from public.transactions t
            left join public.categories c
              on c.id = t.category_id and c.user_id = t.user_id
            where t.user_id = %s
            order by t.occurred_on desc, t.created_at desc
            limit %s offset %s
            """,
            (user_id, limit, offset),
        ).fetchall()
    return list(rows)


@router.get("/transactions/export")
def export_transactions(user_id: UUID = Depends(current_user_id)) -> Response:
    """Export all user transactions in a spreadsheet-friendly CSV format."""
    with get_connection() as connection:
        rows = connection.execute(
            """
            select
              t.description, t.amount, t.direction, t.occurred_on,
              t.status, t.notes, c.name as category_name
            from public.transactions t
            left join public.categories c
              on c.id = t.category_id and c.user_id = t.user_id
            where t.user_id = %s
            order by t.occurred_on asc, t.created_at asc
            """,
            (user_id,),
        ).fetchall()

    output = io.StringIO()
    output.write("\ufeff")
    writer = csv.writer(output, delimiter=";", lineterminator="\r\n")
    writer.writerow(["data", "descricao", "valor", "tipo", "categoria", "status", "observacoes"])
    for row in rows:
        writer.writerow(
            [
                row["occurred_on"].isoformat(),
                safe_csv_text(row["description"]),
                f"{row['amount']:.2f}".replace(".", ","),
                "recebimento" if row["direction"] == Direction.INCOME else "gasto",
                safe_csv_text(row["category_name"]),
                "concluido" if row["status"] == "completed" else "pendente",
                safe_csv_text(row["notes"]),
            ]
        )

    return Response(
        content=output.getvalue(),
        media_type="text/csv",
        headers={"Content-Disposition": 'attachment; filename="cifro-movimentacoes.csv"'},
    )


@router.post("/transactions/import/preview")
async def preview_transaction_import(
    file: UploadFile = File(...),
    user_id: UUID = Depends(current_user_id),
) -> dict:
    """Read an import file without persisting any row."""
    del user_id  # Authentication is required; this first stage is read-only.
    filename = file.filename or "planilha"
    file_type = import_format(filename)
    if file_type is None:
        raise HTTPException(status_code=400, detail="Use um arquivo CSV ou XLSX")

    content = await file.read(IMPORT_MAX_BYTES + 1)
    if not content:
        raise HTTPException(status_code=400, detail="A planilha está vazia")
    if len(content) > IMPORT_MAX_BYTES:
        raise HTTPException(status_code=413, detail="A planilha deve ter no máximo 10 MB")

    try:
        sheets = read_import_sheets(filename, content)
    except ImportLimitError as error:
        raise HTTPException(status_code=413, detail=str(error)) from error
    except (BadZipFile, InvalidFileException, OSError, ParseError, ValueError, KeyError) as error:
        raise HTTPException(status_code=400, detail="Não foi possível ler este arquivo como uma planilha válida") from error
    sheet_previews = []
    total_rows = 0
    valid_rows = 0
    invalid_rows = 0
    ignored_rows = 0
    sheet_types = []

    for sheet_name, rows in sheets:
        header_index, headers, _header_score = find_import_header(rows)
        sheet_type, classification_reason, classification_confidence = classify_import_sheet(
            rows,
            header_index,
            headers,
        )
        sheet_types.append(sheet_type)

        if sheet_type != "transactions":
            sheet_ignored = count_nonempty_import_rows(rows)
            ignored_rows += sheet_ignored
            sheet_previews.append(
                {
                    "name": sheet_name,
                    "type": sheet_type,
                    "header_row": header_index + 1 if header_index is not None else None,
                    "headers": headers,
                    "mapping": detect_import_mapping(headers) if header_index is not None else {field: None for field in IMPORT_FIELDS},
                    "confidence": classification_confidence,
                    "classification": {
                        "type": sheet_type,
                        "reason": classification_reason,
                        "confidence": classification_confidence,
                    },
                    "total_rows": 0,
                    "valid_rows": 0,
                    "invalid_rows": 0,
                    "ignored_rows": sheet_ignored,
                    "rows": [],
                    "errors": [],
                    "importable": False,
                }
            )
            continue

        mapping = detect_import_mapping(headers)
        indexes = import_mapping_indexes(headers, mapping)
        missing_fields = []
        if indexes["date"] is None:
            missing_fields.append("data")
        if indexes["description"] is None:
            missing_fields.append("descrição")
        if indexes["amount"] is None and indexes["income"] is None and indexes["expense"] is None:
            missing_fields.append("valor ou entrada/saída")
        if indexes["direction"] is None and indexes["income"] is None and indexes["expense"] is None:
            missing_fields.append("tipo")

        preview_rows = []
        sheet_total = 0
        sheet_valid = 0
        sheet_invalid = 0
        for row_index, values in enumerate(rows[header_index + 1 :], start=header_index + 2):
            if not any(import_cell_text(value) for value in values):
                continue
            sheet_total += 1
            normalized_row = normalize_import_row(values, indexes, sheet_name, row_index)
            if missing_fields:
                normalized_row["errors"] = [f"Coluna não identificada: {field}" for field in missing_fields]
                normalized_row["valid"] = False
            if normalized_row["valid"]:
                sheet_valid += 1
            else:
                sheet_invalid += 1
            if len(preview_rows) < IMPORT_PREVIEW_LIMIT:
                preview_rows.append(normalized_row)

        total_rows += sheet_total
        valid_rows += sheet_valid
        invalid_rows += sheet_invalid
        sheet_previews.append(
            {
                "name": sheet_name,
                "type": "transactions",
                "header_row": header_index + 1,
                "headers": headers,
                "mapping": mapping,
                "confidence": classification_confidence,
                "classification": {
                    "type": "transactions",
                    "reason": classification_reason,
                    "confidence": classification_confidence,
                },
                "total_rows": sheet_total,
                "valid_rows": sheet_valid,
                "invalid_rows": sheet_invalid,
                "ignored_rows": 0,
                "rows": preview_rows,
                "errors": [],
                "importable": True,
            }
        )

    distinct_types = set(sheet_types)
    if distinct_types == {"transactions"}:
        workbook_type = "transactions"
        workbook_label = "Movimentações"
        workbook_message = "As abas têm estrutura de lançamentos e podem seguir para a próxima etapa."
    elif "transactions" in distinct_types and "summary" in distinct_types:
        workbook_type = "mixed"
        workbook_label = "Planilha mista"
        workbook_message = "Algumas abas têm movimentações; resumos e orçamentos ficarão fora da importação."
    elif "summary" in distinct_types:
        workbook_type = "summary"
        workbook_label = "Resumo mensal ou orçamento"
        workbook_message = "Esta planilha organiza meses, categorias e totais; ela não será convertida em lançamentos automaticamente."
    else:
        workbook_type = "unknown"
        workbook_label = "Formato não reconhecido"
        workbook_message = "Não encontramos uma estrutura segura de movimentações para importar."

    return {
        "filename": filename,
        "format": "xlsx" if filename.lower().endswith(".xlsx") else "csv",
        "workbook_type": workbook_type,
        "workbook_label": workbook_label,
        "workbook_message": workbook_message,
        "total_sheets": len(sheet_previews),
        "total_rows": total_rows,
        "valid_rows": valid_rows,
        "invalid_rows": invalid_rows,
        "ignored_rows": ignored_rows,
        "preview_limit": IMPORT_PREVIEW_LIMIT,
        "sheets": sheet_previews,
    }


@router.post("/transactions", response_model=TransactionRead, status_code=status.HTTP_201_CREATED)
def create_transaction(payload: TransactionCreate, user_id: UUID = Depends(current_user_id)) -> dict:
    with get_connection() as connection:
        category = validate_category_for_direction(
            connection,
            payload.category_id,
            payload.direction.value,
            user_id,
        )
        validate_commitment_for_transaction(
            connection,
            payload.commitment_id,
            payload.direction.value,
            payload.category_id,
            user_id,
        )
        row = connection.execute(
            """
            insert into public.transactions (
              user_id, category_id, commitment_id, description, amount,
              direction, occurred_on, status, notes
            )
            values (%s, %s, %s, %s, %s, %s, %s, %s, %s)
            returning id, description, amount, direction, occurred_on,
              category_id, commitment_id, status, notes, created_at, updated_at
            """,
            (
                user_id,
                payload.category_id,
                payload.commitment_id,
                payload.description.strip(),
                payload.amount,
                payload.direction.value,
                payload.occurred_on,
                payload.status.value,
                payload.notes,
            ),
        ).fetchone()
    row["category_name"] = category["name"] if category else None
    return row


@router.patch("/transactions/{transaction_id}", response_model=TransactionRead)
def update_transaction(
    transaction_id: UUID,
    payload: TransactionUpdate,
    user_id: UUID = Depends(current_user_id),
) -> dict:
    values = payload.model_dump(exclude_unset=True)
    if not values:
        raise HTTPException(status_code=400, detail="No transaction changes provided")

    field_map = {
        "description": "description",
        "amount": "amount",
        "direction": "direction",
        "occurred_on": "occurred_on",
        "category_id": "category_id",
        "commitment_id": "commitment_id",
        "status": "status",
        "notes": "notes",
    }
    assignments = []
    parameters = []
    for field, column in field_map.items():
        if field not in values:
            continue
        value = values[field]
        if field == "description":
            if value is None:
                raise HTTPException(status_code=400, detail="Description cannot be empty")
            value = value.strip()
        elif field in {"direction", "status"}:
            if value is None:
                raise HTTPException(status_code=400, detail=f"{field} cannot be empty")
            value = value.value
        assignments.append(f"{column} = %s")
        parameters.append(value)

    with get_connection() as connection:
        current = connection.execute(
            """
            select id, direction, category_id, commitment_id
            from public.transactions
            where id = %s and user_id = %s
            for update
            """,
            (transaction_id, user_id),
        ).fetchone()
        if not current:
            raise HTTPException(status_code=404, detail="Transaction not found")

        effective_direction = values.get("direction", current["direction"])
        if isinstance(effective_direction, Direction):
            effective_direction = effective_direction.value
        effective_category_id = values.get("category_id", current["category_id"])
        effective_commitment_id = values.get("commitment_id", current["commitment_id"])

        category = None
        if effective_category_id and ({"category_id", "direction"} & values.keys()):
            category = validate_category_for_direction(
                connection,
                effective_category_id,
                effective_direction,
                user_id,
            )
        elif effective_category_id:
            category = connection.execute(
                """
                select name
                from public.categories
                where id = %s and user_id = %s
                """,
                (effective_category_id, user_id),
            ).fetchone()

        if effective_commitment_id and ({"commitment_id", "category_id", "direction"} & values.keys()):
            validate_commitment_for_transaction(
                connection,
                effective_commitment_id,
                effective_direction,
                effective_category_id,
                user_id,
            )

        parameters.extend([transaction_id, user_id])
        row = connection.execute(
            f"""
            update public.transactions
            set {', '.join(assignments)}
            where id = %s and user_id = %s
            returning id, description, amount, direction, occurred_on,
              category_id, commitment_id, status, notes, created_at, updated_at
            """,
            parameters,
        ).fetchone()

        if not row:
            raise HTTPException(status_code=404, detail="Transaction not found")

        category_name = category["name"] if category else None
        if row["category_id"] and category_name is None:
            category = connection.execute(
                """
                select name
                from public.categories
                where id = %s and user_id = %s
                """,
                (row["category_id"], user_id),
            ).fetchone()
            category_name = category["name"] if category else None

    row["category_name"] = category_name
    return row


@router.delete("/transactions/{transaction_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_transaction(transaction_id: UUID, user_id: UUID = Depends(current_user_id)) -> None:
    with get_connection() as connection:
        row = connection.execute(
            """
            delete from public.transactions
            where id = %s and user_id = %s
            returning id
            """,
            (transaction_id, user_id),
        ).fetchone()

    if not row:
        raise HTTPException(status_code=404, detail="Transaction not found")


@router.get("/dashboard", response_model=DashboardRead)
def dashboard(
    year: Annotated[int, Query(ge=2000, le=2100)] | None = None,
    month: Annotated[int, Query(ge=1, le=12)] | None = None,
    user_id: UUID = Depends(current_user_id),
) -> dict:
    today = date.today()
    selected_year = year or today.year
    selected_month = month or today.month
    current_start, current_end = month_bounds(selected_year, selected_month)
    next_year, next_month_number = next_month(selected_year, selected_month)
    next_start, next_end = month_bounds(next_year, next_month_number)
    with get_connection() as connection:
        current_totals = connection.execute(
            """
            select direction, coalesce(sum(amount), 0) as total
            from public.transactions
            where user_id = %s and status = 'completed'
              and occurred_on >= %s and occurred_on < %s
            group by direction
            """,
            (user_id, current_start, current_end),
        ).fetchall()

        budget_summary = build_budget_summary(
            connection,
            user_id,
            selected_year,
            selected_month,
        )

        planned_totals = connection.execute(
            """
            select direction, coalesce(sum(amount), 0) as total
            from public.transactions
            where user_id = %s and status = 'planned'
              and occurred_on >= %s and occurred_on < %s
            group by direction
            """,
            (user_id, next_start, next_end),
        ).fetchall()

        commitment_rows = connection.execute(
            """
            select
              c.id, c.name, c.amount, c.direction, c.commitment_type,
              c.frequency, c.due_rule, c.due_day, c.due_month, c.business_day_number,
              c.starts_on, c.next_due_on, c.ends_on,
              cat.name as category_name
            from public.commitments c
            left join public.categories cat
              on cat.id = c.category_id and cat.user_id = c.user_id
            where c.user_id = %s and c.is_active = true
            order by c.next_due_on asc, c.name asc
            """,
            (user_id,),
        ).fetchall()

        recent = connection.execute(
            """
            select
              t.id, t.description, t.amount, t.direction, t.occurred_on,
              t.category_id, t.commitment_id, t.status, t.notes,
              t.created_at, t.updated_at, c.name as category_name
            from public.transactions t
            left join public.categories c
              on c.id = t.category_id and c.user_id = t.user_id
            where t.user_id = %s
            order by t.occurred_on desc, t.created_at desc
            limit 6
            """,
            (user_id,),
        ).fetchall()

    commitments = []
    for row in commitment_rows:
        projected_date = projected_commitment_date(row, next_year, next_month_number)
        if projected_date is None:
            continue
        projected = dict(row)
        projected["next_due_on"] = projected_date
        commitments.append(projected)
    commitments.sort(key=lambda row: (row["next_due_on"], row["name"].lower()))

    current = {row["direction"]: as_money(row["total"]) for row in current_totals}
    planned = {row["direction"]: as_money(row["total"]) for row in planned_totals}
    commitment_income = sum(
        (row["amount"] for row in commitments if row["direction"] == Direction.INCOME),
        Decimal("0.00"),
    )
    commitment_expenses = sum(
        (row["amount"] for row in commitments if row["direction"] == Direction.EXPENSE),
        Decimal("0.00"),
    )
    next_income = planned.get(Direction.INCOME, Decimal("0.00")) + commitment_income
    next_expenses = planned.get(Direction.EXPENSE, Decimal("0.00")) + commitment_expenses
    budget_overview = BudgetDashboardRead(
        base_amount=budget_summary["base_amount"],
        allocated_amount=budget_summary["allocated_amount"],
        total_percentage=budget_summary["total_percentage"],
        unallocated_percentage=budget_summary["unallocated_percentage"],
        unallocated_amount=budget_summary["unallocated_amount"],
        allocation_count=len(budget_summary["allocations"]),
    )

    return {
        "month": f"{selected_year:04d}-{selected_month:02d}",
        "next_month": f"{next_year:04d}-{next_month_number:02d}",
        "current": DashboardPeriod(
            income=current.get(Direction.INCOME, Decimal("0.00")),
            expenses=current.get(Direction.EXPENSE, Decimal("0.00")),
            available=current.get(Direction.INCOME, Decimal("0.00"))
            - current.get(Direction.EXPENSE, Decimal("0.00")),
        ),
        "next_month_summary": DashboardPeriod(
            income=next_income,
            expenses=next_expenses,
            available=next_income - next_expenses,
        ),
        "next_month_commitments": [CommitmentPreview(**row) for row in commitments],
        "recent_transactions": [TransactionRead(**row) for row in recent],
        "budget": budget_overview,
    }


app.include_router(router)
